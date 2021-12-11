from openpyxl import load_workbook
from datetime import date,datetime
import tkinter as tk
from tkinter.filedialog import askopenfile

def click():
    global end, min_time
    end = tb1.get()
    min_time = int(tb2.get())
    root.destroy()

root = tk.Tk()
root.geometry("400x350")

def fn1():
    global ms_attend
    ms_attend = askopenfile().name
    bt1.configure(text="File chosen")

def fn2():
    global attend_sheet
    attend_sheet = askopenfile().name
    bt2.configure(text="File chosen")

lb1 = tk.Label(root, text="Attendance bot")
lb1.grid(row=0, column=0, pady=10)

lb2 = tk.Label(root, text="Enter the time at which class ends\n (HH:MM:SS) in 24 hr format")
lb2.grid(row=1, column=0, pady=10)

lb3 = tk.Label(root, text="Enter the minimum time (in min)\n required to be in the meeting")
lb3.grid(row=2, column=0, pady=10)

lb4 = tk.Label(root, text="Choose the excel file downloaded from MS teams")
lb4.grid(row=3, column=0, pady=10)

lb5 = tk.Label(root, text="Choose the attendance sheet")
lb5.grid(row=4, column=0, pady=10)

tb1 = tk.Entry(root)
tb1.grid(row=1, column=1)

tb2 = tk.Entry(root)
tb2.grid(row=2, column=1)

bt1 = tk.Button(root, text="Choose file", command=fn1)
bt1.grid(row=3, column=1)

bt2 = tk.Button(root, text="Choose file", command=fn2)
bt2.grid(row=4, column=1)

submit = tk.Button(root, text="Submit", command=click)
submit.grid(row=5, column=0, pady=10)

lb6 = tk.Label(root, text="   #Developed by Harsiddak Singh Bedi",font=("Times New Roman", 12, "bold"))
lb6.grid(row=6, column=0)

root.mainloop()

end = datetime.strptime(end,"%H:%M:%S")

D = {}

wb1 = load_workbook(ms_attend)
sheet1 = wb1.active

wb2 = load_workbook(attend_sheet)
sheet2 = wb2.active

for row in range(5):
    for col in range(5):
        x = list(sheet2.rows)[row][col].value
        if x != None:
            if "NAME" in str(x).upper():
                head_row = row + 1
                head_col = col
                break

for i in range(len(sheet2[head_row])):
    if sheet2[head_row][i].value == None:
        col_num = i
        break

sheet2[head_row][col_num].value = date.today().strftime("%d %B %Y (%A)").upper()

iter_row = iter(sheet1.rows)
next(iter_row)

for row in iter_row:
    if row[0].value != None:
        name = row[0].value.strip().upper()
        s = datetime.strptime(row[2].value.split(" ")[1],"%H:%M:%S")
        if name in D:
            if row[1].value == "Left":
                D[name] -= end - s
            else:
                D[name] += end - s
        else:
            D[name] = end - s
               
for row in sheet2.rows:
    if row[head_col].value != None:
        x = row[head_col].value.strip().upper()
        if x in D:
            if int(D[x].seconds) >= min_time*60:
                row[col_num].value = "P"

wb2.save(attend_sheet)